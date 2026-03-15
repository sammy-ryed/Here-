"""
Flask Backend for Face Recognition Attendance System
Handles face registration and attendance processing using RetinaFace and ArcFace
OPTIMIZED: GPU support + Batch processing + Threading + Caching
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import numpy as np
from datetime import datetime
import logging
from werkzeug.utils import secure_filename
import base64
from io import BytesIO
from PIL import Image
import cv2
from concurrent.futures import ThreadPoolExecutor
import threading

from utils.face_detector import FaceDetector
from utils.face_recognizer import FaceRecognizer
from db.database import Database

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = 'uploads'
UNRECOGNIZED_FOLDER = 'unrecognized'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
CONFIDENCE_THRESHOLD = 0.769  # Cosine similarity threshold

# Create necessary folders
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(UNRECOGNIZED_FOLDER, exist_ok=True)

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize components
face_detector = FaceDetector()
face_recognizer = FaceRecognizer()
database = Database()

# Thread pool for parallel processing (2 workers to avoid overwhelming DeepFace)
# Note: DeepFace has threading issues, so we limit to 2 concurrent workers
executor = ThreadPoolExecutor(max_workers=2)

def allowed_file(filename):
    """Check if file has an allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_single_face_for_attendance(face_idx, face, filepath, all_students, recognized_student_ids_lock, recognized_student_ids):
    """
    Process a single face for attendance (thread-safe).
    Uses vectorised matrix search via find_match for fast, accurate matching.
    Returns: (recognized_student, unrecognized_face_data) — one will be None
    """
    try:
        logger.info(f"Processing Face #{face_idx + 1}...")

        embedding = face_recognizer.get_embedding(filepath, face)
        if embedding is None:
            logger.info(f"  ❌ Face #{face_idx + 1}: failed to generate embedding")
            return None, None

        # Filter out already-marked students for this image
        with recognized_student_ids_lock:
            remaining_students = [s for s in all_students if s['id'] not in recognized_student_ids]

        if not remaining_students:
            return None, None

        # Matrix search — single BLAS multiply across all remaining students
        result = face_recognizer.find_match(
            embedding, remaining_students, threshold=CONFIDENCE_THRESHOLD, db=database
        )

        if result is not None:
            student_id, confidence = result
            student = next((s for s in remaining_students if s['id'] == student_id), None)
            if student is None:
                return None, None

            with recognized_student_ids_lock:
                if student_id in recognized_student_ids:
                    logger.info(f"  ⚠️ Face #{face_idx + 1}: already marked by another thread")
                    return None, None
                recognized_student_ids.add(student_id)

            logger.info(f"  ✅ Face #{face_idx + 1}: MATCH — {student['name']} (confidence: {confidence:.4f})")
            return {
                'id': student['id'],
                'name': student['name'],
                'roll_no': student['roll_no'],
                'confidence': float(confidence)
            }, None
        else:
            logger.info(f"  ❌ Face #{face_idx + 1}: no match (below threshold {CONFIDENCE_THRESHOLD})")
            return None, {'bbox': face['bbox'], 'embedding': embedding}

    except Exception as e:
        logger.error(f"Error processing face {face_idx}: {str(e)}")
        return None, None

def process_single_image_for_registration(image_file, roll_no, idx):
    """
    Process a single image for registration (thread-safe)
    Returns: (embedding, error_message)
    """
    try:
        if not image_file or not allowed_file(image_file.filename):
            return None, f"Invalid file: {image_file.filename}"
        
        # Save image temporarily
        filename = f"{roll_no}_{idx}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{threading.get_ident()}.jpg"
        filepath = save_image(image_file, filename)
        
        try:
            # Detect faces
            faces = face_detector.detect_faces(filepath)
            
            if len(faces) == 0:
                return None, "No face detected"
            
            if len(faces) > 1:
                logger.warning(f"Multiple faces in image {idx + 1}, using largest")
            
            # Get embedding from largest face
            largest_face = max(faces, key=lambda x: (x['bbox'][2] - x['bbox'][0]) * (x['bbox'][3] - x['bbox'][1]))
            embedding = face_recognizer.get_embedding(filepath, largest_face)
            
            return embedding, None
        finally:
            # Clean up temporary file
            if os.path.exists(filepath):
                os.remove(filepath)
    
    except Exception as e:
        logger.error(f"Error processing image {idx}: {str(e)}")
        return None, str(e)

def save_image(file_data, filename):
    """Save uploaded image file"""
    filepath = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
    
    if isinstance(file_data, str):
        # Base64 encoded image
        image_data = base64.b64decode(file_data.split(',')[1] if ',' in file_data else file_data)
        with open(filepath, 'wb') as f:
            f.write(image_data)
    else:
        # File object
        file_data.save(filepath)
    
    return filepath

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'services': {
            'face_detector': face_detector.is_ready(),
            'face_recognizer': face_recognizer.is_ready(),
            'database': database.is_connected()
        }
    })

@app.route('/register_face', methods=['POST'])
def register_face():
    """
    Register a new student with face images
    OPTIMIZED: Batch processing with threading for faster registration
    Expected: name, roll_no, and multiple images
    """
    try:
        # Get form data
        name = request.form.get('name')
        roll_no = request.form.get('roll_no')
        
        if not name or not roll_no:
            return jsonify({'error': 'Name and roll number are required'}), 400
        
        # Get uploaded images
        images = request.files.getlist('images')
        
        if not images or len(images) == 0:
            return jsonify({'error': 'At least one image is required'}), 400
        
        logger.info(f"🚀 BATCH PROCESSING: Registering {name} with {len(images)} images using {executor._max_workers} threads")
        
        start_time = datetime.now()
        
        # Process images in parallel using ThreadPoolExecutor
        futures = []
        for idx, image_file in enumerate(images):
            future = executor.submit(process_single_image_for_registration, image_file, roll_no, idx)
            futures.append(future)
        
        # Collect results
        all_embeddings = []
        processed_count = 0
        error_count = 0
        
        for idx, future in enumerate(futures):
            embedding, error = future.result()
            
            if embedding is not None:
                all_embeddings.append(embedding)
                processed_count += 1
            else:
                error_count += 1
                logger.warning(f"Image {idx + 1} failed: {error}")
        
        if len(all_embeddings) == 0:
            return jsonify({'error': 'No valid face embeddings could be extracted'}), 400
        
        # Average embeddings for better accuracy
        avg_embedding = np.mean(all_embeddings, axis=0)
        
        # Store in database
        student_id = database.add_student(name, roll_no, avg_embedding)
        
        # Store individual embeddings as well for better matching
        for embedding in all_embeddings:
            database.add_embedding(student_id, embedding)
        
        elapsed_time = (datetime.now() - start_time).total_seconds()
        
        logger.info(f"✅ Successfully registered {name} with {processed_count}/{len(images)} samples in {elapsed_time:.2f}s")
        
        return jsonify({
            'success': True,
            'student_id': student_id,
            'name': name,
            'roll_no': roll_no,
            'samples_processed': processed_count,
            'samples_failed': error_count,
            'processing_time': f"{elapsed_time:.2f}s"
        }), 201
    
    except Exception as e:
        logger.error(f"Error in register_face: {str(e)}", exc_info=True)
        return jsonify({'error': f'Registration failed: {str(e)}'}), 500

@app.route('/process_attendance', methods=['POST'])
def process_attendance():
    """
    Process attendance from classroom photo
    Detects all faces and matches against registered students
    """
    try:
        # Get uploaded image
        if 'image' not in request.files:
            return jsonify({'error': 'No image provided'}), 400
        
        image_file = request.files['image']
        class_name = request.form.get('class_name', 'default')
        
        if not image_file or not allowed_file(image_file.filename):
            return jsonify({'error': 'Invalid image file'}), 400
        
        logger.info(f"Processing attendance for class: {class_name}")
        
        # Save image temporarily
        filename = f"attendance_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
        filepath = save_image(image_file, filename)
        
        # Detect all faces in the image
        faces = face_detector.detect_faces(filepath)
        logger.info(f"Detected {len(faces)} faces in the image")
        
        if len(faces) == 0:
            os.remove(filepath)
            return jsonify({
                'present': [],
                'unrecognized': [],
                'total_faces': 0,
                'message': 'No faces detected in the image'
            })
        
        # Get all registered students
        all_students = database.get_all_students()
        
        logger.info(f"\n{'='*60}")
        logger.info(f"� BATCH PROCESSING: {len(faces)} FACES using {executor._max_workers} threads")
        logger.info(f"{'='*60}\n")
        
        start_time = datetime.now()
        
        # Thread-safe set for tracking recognized students
        recognized_student_ids = set()
        recognized_student_ids_lock = threading.Lock()
        
        # Process faces in parallel using ThreadPoolExecutor
        futures = []
        for idx, face in enumerate(faces):
            future = executor.submit(
                process_single_face_for_attendance,
                idx, face, filepath, all_students,
                recognized_student_ids_lock, recognized_student_ids
            )
            futures.append(future)
        
        # Collect results
        recognized_students = []
        unrecognized_faces = []
        
        for idx, future in enumerate(futures):
            try:
                # Add timeout to prevent hanging indefinitely
                recognized_student, unrecognized_face = future.result(timeout=60)  # 60 second timeout per face
                
                if recognized_student:
                    recognized_students.append(recognized_student)
                elif unrecognized_face:
                    unrecognized_faces.append(unrecognized_face)
            except TimeoutError:
                logger.error(f"⏱️ Face processing thread {idx} timed out after 60 seconds")
            except Exception as e:
                logger.error(f"Error getting result from face processing thread {idx}: {str(e)}", exc_info=True)
        
        # Save unrecognized faces for review
        img = cv2.imread(filepath)
        unrecognized_images = []
        
        for idx, face_data in enumerate(unrecognized_faces):
            bbox = face_data['bbox']
            face_img = Image.open(filepath)
            face_crop = face_img.crop((bbox[0], bbox[1], bbox[2], bbox[3]))
            
            unrecognized_filename = f"unrecognized_{datetime.now().strftime('%Y%m%d%H%M%S')}_{idx}.jpg"
            unrecognized_path = os.path.join(UNRECOGNIZED_FOLDER, unrecognized_filename)
            face_crop.save(unrecognized_path)
            
            unrecognized_images.append(unrecognized_filename)
        
        elapsed_time = (datetime.now() - start_time).total_seconds()
        
        logger.info(f"{'='*60}")
        logger.info(f"✅ RESULTS: {len(recognized_students)} students marked PRESENT in {elapsed_time:.2f}s")
        logger.info(f"{'='*60}\n")
        
        # Clean up temporary file
        os.remove(filepath)
        
        # Mark attendance for recognized students as present (with confidence)
        today = datetime.now().strftime('%Y-%m-%d')
        for student in recognized_students:
            database.mark_attendance(student['id'], today, 'present',
                                     confidence=student.get('confidence'))
        
        # Get today's attendance to check who is already marked present
        today_attendance = database.get_attendance_by_date(today)
        present_ids = [record['id'] for record in today_attendance if record.get('status') == 'present']
        
        # Get list of absent students (registered but not recognized AND not already marked present)
        recognized_ids = [s['id'] for s in recognized_students]
        absent_students = [s for s in all_students if s['id'] not in recognized_ids and s['id'] not in present_ids]
        
        # Mark absent students as absent in the database (only if not already present)
        for student in absent_students:
            database.mark_attendance(student['id'], today, 'absent')
        
        return jsonify({
            'present': recognized_students,
            'absent': [{'id': s['id'], 'name': s['name'], 'roll_no': s['roll_no']} for s in absent_students],
            'unrecognized': unrecognized_images,
            'total_faces': len(faces),
            'processing_time': f"{elapsed_time:.2f}s",
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        logger.error(f"Error in process_attendance: {str(e)}", exc_info=True)
        return jsonify({'error': f'Attendance processing failed: {str(e)}'}), 500

@app.route('/process_attendance_batch', methods=['POST'])
def process_attendance_batch():
    """
    Process attendance from multiple classroom photos
    Detects all faces across all photos and matches against registered students
    Uses threading for faster processing
    """
    try:
        # Get uploaded images
        if 'images' not in request.files:
            return jsonify({'error': 'No images provided'}), 400
        
        image_files = request.files.getlist('images')
        class_name = request.form.get('class_name', 'default')
        
        if not image_files or len(image_files) == 0:
            return jsonify({'error': 'No images provided'}), 400
        
        logger.info(f"🚀 BATCH PROCESSING: {len(image_files)} photos for class: {class_name}")
        
        # Get all registered students
        all_students = database.get_all_students()
        
        # Thread-safe tracking
        recognized_student_ids = set()
        recognized_student_ids_lock = threading.Lock()
        all_recognized_students = []
        all_unrecognized_faces = []
        total_faces_count = 0
        
        # Process each photo
        for photo_idx, image_file in enumerate(image_files):
            if not allowed_file(image_file.filename):
                logger.warning(f"Skipping invalid file: {image_file.filename}")
                continue
            
            try:
                # Save image temporarily
                filename = f"attendance_batch_{datetime.now().strftime('%Y%m%d%H%M%S')}_{photo_idx}.jpg"
                filepath = save_image(image_file, filename)
                
                # Detect all faces in this image
                faces = face_detector.detect_faces(filepath)
                logger.info(f"📸 Photo {photo_idx + 1}/{len(image_files)}: Detected {len(faces)} faces")
                total_faces_count += len(faces)
                
                if len(faces) > 0:
                    # Process faces in parallel
                    futures = []
                    for idx, face in enumerate(faces):
                        future = executor.submit(
                            process_single_face_for_attendance,
                            idx, face, filepath, all_students,
                            recognized_student_ids_lock, recognized_student_ids
                        )
                        futures.append(future)
                    
                    # Collect results
                    for future in futures:
                        try:
                            recognized_student, unrecognized_face = future.result(timeout=60)
                            
                            if recognized_student:
                                all_recognized_students.append(recognized_student)
                            elif unrecognized_face:
                                # Save unrecognized face image
                                bbox = unrecognized_face['bbox']
                                img = cv2.imread(filepath)
                                x1, y1, x2, y2 = [int(coord) for coord in bbox]
                                face_crop = img[y1:y2, x1:x2]
                                
                                unrecognized_filename = f"unrecognized_{datetime.now().strftime('%Y%m%d%H%M%S')}_{photo_idx}.jpg"
                                unrecognized_path = os.path.join(UNRECOGNIZED_FOLDER, unrecognized_filename)
                                cv2.imwrite(unrecognized_path, face_crop)
                                
                                all_unrecognized_faces.append(unrecognized_filename)
                        
                        except Exception as e:
                            logger.error(f"Error processing face: {str(e)}")
                
                # Clean up temporary file
                os.remove(filepath)
                
            except Exception as e:
                logger.error(f"Error processing photo {image_file.filename}: {str(e)}")
                continue
        
        # Mark attendance for recognized students as present (with confidence)
        today = datetime.now().strftime('%Y-%m-%d')
        for student in all_recognized_students:
            database.mark_attendance(student['id'], today, 'present',
                                     confidence=student.get('confidence'))
        
        # Get today's attendance to check who is already marked present
        today_attendance = database.get_attendance_by_date(today)
        present_ids = [record['id'] for record in today_attendance if record.get('status') == 'present']
        
        # Get list of absent students (registered but not recognized AND not already marked present)
        recognized_ids = [s['id'] for s in all_recognized_students]
        absent_students = [s for s in all_students if s['id'] not in recognized_ids and s['id'] not in present_ids]
        
        # Mark absent students as absent (database now prevents overwriting present with absent)
        for student in absent_students:
            database.mark_attendance(student['id'], today, 'absent')
        
        logger.info(f"✅ Batch complete: {len(image_files)} photos, {total_faces_count} faces, {len(all_recognized_students)} students recognized")
        
        return jsonify({
            'present': all_recognized_students,
            'absent': [{'id': s['id'], 'name': s['name'], 'roll_no': s['roll_no']} for s in absent_students],
            'unrecognized': all_unrecognized_faces,
            'total_faces': total_faces_count,
            'photos_processed': len(image_files),
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        logger.error(f"Error in process_attendance_batch: {str(e)}", exc_info=True)
        return jsonify({'error': f'Batch attendance processing failed: {str(e)}'}), 500

@app.route('/unrecognized/<filename>', methods=['GET'])
def get_unrecognized_face(filename):
    """Serve unrecognized face image"""
    try:
        filepath = os.path.join(UNRECOGNIZED_FOLDER, secure_filename(filename))
        if os.path.exists(filepath):
            return send_file(filepath, mimetype='image/jpeg')
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        logger.error(f"Error serving unrecognized face: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/students', methods=['GET'])
def get_all_students():
    """Get all registered students with full profile fields"""
    try:
        students = database.get_all_students_full()
        return jsonify({
            'students': students,
            'total': len(students)
        })
    except Exception as e:
        logger.error(f"Error getting students: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/attendance/report', methods=['GET'])
def get_attendance_report():
    """Get attendance report for a specific date or date range"""
    try:
        start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-%d'))
        end_date = request.args.get('end_date', start_date)
        
        report = database.get_attendance_report(start_date, end_date)
        
        return jsonify({
            'report': report,
            'start_date': start_date,
            'end_date': end_date
        })
    except Exception as e:
        logger.error(f"Error generating attendance report: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    """Get dashboard statistics including today's attendance"""
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Get total students
        all_students_data = database.get_all_students()
        total_students = len(all_students_data)
        
        # Get today's attendance
        today_attendance = database.get_attendance_by_date(today)
        
        # Create a map of student_id to attendance status + confidence
        attendance_map = {}
        for record in today_attendance:
            if record.get('status'):
                attendance_map[record['id']] = {
                    'status': record['status'],
                    'confidence': record.get('confidence')
                }

        # Convert students to JSON-serializable format with attendance status
        all_students = []
        for student in all_students_data:
            student_id = student['id']
            att = attendance_map.get(student_id, {})
            attendance_status = att.get('status', 'N/A')
            attendance_confidence = att.get('confidence')  # None for manual marks

            all_students.append({
                'id': student_id,
                'name': student['name'],
                'roll_no': student['roll_no'],
                'attendance_status': attendance_status,
                'attendance_confidence': attendance_confidence
            })
        
        # Count present and absent
        present_today = sum(1 for record in today_attendance if record.get('status') == 'present')
        absent_today = total_students - present_today if present_today > 0 else 0
        
        # Calculate attendance rate
        attendance_rate = (present_today / total_students * 100) if total_students > 0 else 0
        
        return jsonify({
            'total_students': total_students,
            'present_today': present_today,
            'absent_today': absent_today,
            'attendance_rate': round(attendance_rate, 1),
            'date': today,
            'students': all_students
        })
    except Exception as e:
        logger.error(f"Error getting dashboard stats: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/students/bulk-import', methods=['POST'])
def bulk_import_students():
    """
    Bulk-import students from an Excel file.
    AUTO-DETECTS MODE:
      - If no 'Drive Link' column → SELF-REGISTRATION MODE:
          Creates pending student records + secure one-time tokens.
          Faculty distributes the links to students privately.
      - If 'Drive Link' column present → DRIVE MODE (legacy):
          Downloads Drive folders and registers immediately.
    Required columns: Name, Roll No  (+ optional: Section, Course, Department, Room No)
    """
    import tempfile
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'openpyxl is not installed. Run: pip install openpyxl'}), 500

    if 'file' not in request.files:
        return jsonify({'error': 'No Excel file provided'}), 400

    excel_file = request.files['file']
    if not excel_file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only .xlsx or .xls files are accepted'}), 400

    tmp_excel = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    excel_file.save(tmp_excel.name)
    tmp_excel.close()

    try:
        wb = openpyxl.load_workbook(tmp_excel.name)
        ws = wb.active

        raw_headers = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in ws[1]]
        col_map = {}
        ALIASES = {
            'name':       ('name', 'student name', 'full name', 'student_name'),
            'roll_no':    ('roll no', 'roll_no', 'roll number', 'rollno', 'roll'),
            'section':    ('section',),
            'course':     ('course', 'programme', 'program'),
            'dept':       ('department', 'dept'),
            'room_no':    ('room no', 'room_no', 'room', 'room number'),
            'drive_link': ('drive link', 'drive_link', 'photos link', 'link', 'photo link', 'photos', 'folder link'),
        }
        for key, aliases in ALIASES.items():
            for i, h in enumerate(raw_headers):
                if h in aliases:
                    col_map[key] = i
                    break

        missing_required = [k for k in ('name', 'roll_no') if k not in col_map]
        if missing_required:
            return jsonify({'error': f"Excel is missing required column(s): {', '.join(missing_required)}. "
                                     f"Required: Name, Roll No"}), 400

        token_mode = 'drive_link' not in col_map
        logger.info(f"[bulk-import] Mode: {'SELF-REGISTRATION TOKENS' if token_mode else 'DRIVE LINKS'}")

        # ── TOKEN MODE ──────────────────────────────────────────────────────────
        if token_mode:
            tokens = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                def cell(key):
                    idx = col_map.get(key)
                    return str(row[idx]).strip() if idx is not None and row[idx] is not None else ''

                name    = cell('name')
                roll_no = cell('roll_no')
                section = cell('section') or None
                course  = cell('course') or None
                dept    = cell('dept') or None
                room_no = cell('room_no') or None

                if not name or not roll_no:
                    tokens.append({'roll_no': roll_no or f'row_{row_idx}', 'name': name,
                                   'status': 'skipped', 'message': 'Missing name or roll_no'})
                    continue

                existing = database.get_student_by_roll_no(roll_no)
                if existing:
                    tokens.append({'roll_no': roll_no, 'name': name,
                                   'status': 'skipped', 'message': 'Already in system'})
                    continue

                try:
                    student_id = database.add_pending_student(
                        name, roll_no, section=section, course=course, dept=dept, room_no=room_no
                    )
                    token = database.create_registration_token(student_id, expires_days=7)
                    tokens.append({
                        'roll_no': roll_no, 'name': name,
                        'status': 'pending', 'token': token,
                        'message': 'Awaiting student self-registration'
                    })
                    logger.info(f"[bulk-import] Created token for {name} ({roll_no})")
                except ValueError as e:
                    tokens.append({'roll_no': roll_no, 'name': name, 'status': 'failed', 'message': str(e)})

            pending_count = sum(1 for t in tokens if t['status'] == 'pending')
            skipped_count = sum(1 for t in tokens if t['status'] == 'skipped')
            failed_count  = sum(1 for t in tokens if t['status'] == 'failed')

            return jsonify({
                'success': True,
                'mode': 'tokens',
                'total':   len(tokens),
                'pending': pending_count,
                'skipped': skipped_count,
                'failed':  failed_count,
                'tokens':  tokens
            })

        # ── DRIVE MODE (legacy) ───────────────────────────────────────────────
        BLANK_LINK_VALUES = {'', 'na', 'n/a', 'none', 'nil', '-', '--', 'n.a.', 'n.a', 'null'}

        def is_blank_link(val: str) -> bool:
            return val.lower().strip() in BLANK_LINK_VALUES

        try:
            import gdown
        except ImportError:
            return jsonify({'error': 'gdown is not installed. Run: pip install gdown'}), 500

        results = []
        token_results = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            def cell(key):
                idx = col_map.get(key)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else ''

            name       = cell('name')
            roll_no    = cell('roll_no')
            drive_link = cell('drive_link')
            section    = cell('section') or None
            course     = cell('course') or None
            dept       = cell('dept') or None
            room_no    = cell('room_no') or None

            if not name or not roll_no:
                results.append({'roll_no': roll_no or f'row_{row_idx}', 'name': name,
                                 'status': 'skipped', 'message': 'Missing name or roll_no'})
                continue

            # No drive link provided — fall back to token registration
            if is_blank_link(drive_link):
                if database.get_student_by_roll_no(roll_no):
                    token_results.append({'roll_no': roll_no, 'name': name,
                                          'status': 'skipped', 'message': 'Already in system'})
                    continue
                try:
                    student_id = database.add_pending_student(
                        name, roll_no, section=section, course=course, dept=dept, room_no=room_no
                    )
                    token = database.create_registration_token(student_id, expires_days=7)
                    token_results.append({
                        'roll_no': roll_no, 'name': name,
                        'status': 'pending', 'token': token,
                        'message': 'No drive link — awaiting self-registration'
                    })
                    logger.info(f"[bulk-import] No drive link for {name} ({roll_no}), generated token")
                except ValueError as e:
                    token_results.append({'roll_no': roll_no, 'name': name,
                                          'status': 'failed', 'message': str(e)})
                continue

            if not drive_link:
                results.append({'roll_no': roll_no or f'row_{row_idx}', 'name': name,
                                 'status': 'skipped', 'message': 'Missing drive_link'})
                continue

            if database.get_student_by_roll_no(roll_no):
                results.append({'roll_no': roll_no, 'name': name,
                                 'status': 'skipped', 'message': 'Already registered'})
                continue

            with tempfile.TemporaryDirectory() as tmp_dir:
                try:
                    downloaded = gdown.download_folder(drive_link, output=tmp_dir, quiet=True, use_cookies=False)
                    if downloaded is None or len(downloaded) == 0:
                        results.append({'roll_no': roll_no, 'name': name, 'status': 'failed',
                                         'message': 'Could not download Drive folder — set sharing to "Anyone with the link"'})
                        continue
                except Exception as e:
                    results.append({'roll_no': roll_no, 'name': name, 'status': 'failed',
                                     'message': f'Drive download error: {str(e)}'})
                    continue

                image_exts = {'.jpg', '.jpeg', '.png', '.webp', '.bmp'}
                image_paths = []
                for root, _, files in os.walk(tmp_dir):
                    for f in sorted(files):
                        if os.path.splitext(f)[1].lower() in image_exts:
                            image_paths.append(os.path.join(root, f))

                if not image_paths:
                    results.append({'roll_no': roll_no, 'name': name, 'status': 'failed',
                                     'message': 'No image files found in Drive folder'})
                    continue

                all_embeddings = []
                for img_path in image_paths[:8]:
                    try:
                        faces = face_detector.detect_faces(img_path)
                        if not faces:
                            continue
                        largest = max(faces, key=lambda f: (f['bbox'][2] - f['bbox'][0]) * (f['bbox'][3] - f['bbox'][1]))
                        emb = face_recognizer.get_embedding(img_path, largest)
                        if emb is not None:
                            all_embeddings.append(emb)
                    except Exception as e:
                        logger.warning(f"[bulk-import] {name} — image error: {e}")

                if not all_embeddings:
                    results.append({'roll_no': roll_no, 'name': name, 'status': 'failed',
                                     'message': 'No usable face found in any downloaded photo'})
                    continue

                avg_embedding = np.mean(all_embeddings, axis=0)
                try:
                    student_id = database.add_student(
                        name, roll_no, avg_embedding,
                        section=section, course=course, dept=dept, room_no=room_no
                    )
                    for emb in all_embeddings:
                        database.add_embedding(student_id, emb)
                    results.append({
                        'roll_no': roll_no, 'name': name, 'status': 'success',
                        'message': f'Registered with {len(all_embeddings)} photo(s)',
                        'student_id': student_id
                    })
                except ValueError as e:
                    results.append({'roll_no': roll_no, 'name': name, 'status': 'failed', 'message': str(e)})

        success_count = sum(1 for r in results if r['status'] == 'success')
        failed_count  = sum(1 for r in results if r['status'] == 'failed')
        skipped_count = sum(1 for r in results if r['status'] == 'skipped')
        token_pending = sum(1 for t in token_results if t['status'] == 'pending')
        token_skipped = sum(1 for t in token_results if t['status'] == 'skipped')
        token_failed  = sum(1 for t in token_results if t['status'] == 'failed')
        logger.info(f"[bulk-import][drive] Done — {success_count} registered, {failed_count} failed, {skipped_count} skipped, {token_pending} tokens")

        mode = 'mixed' if token_results else 'drive'
        return jsonify({
            'success': True,
            'mode': mode,
            'total':      len(results) + len(token_results),
            'registered': success_count,
            'failed':     failed_count + token_failed,
            'skipped':    skipped_count + token_skipped,
            'results':    results,
            # token portion (only present in mixed mode)
            'pending':    token_pending,
            'tokens':     token_results if token_results else None,
        })

    except Exception as e:
        logger.error(f"[bulk-import] Unexpected error: {str(e)}", exc_info=True)
        return jsonify({'error': f'Bulk import failed: {str(e)}'}), 500
    finally:
        os.unlink(tmp_excel.name)


# ── SELF-REGISTRATION ROUTES ─────────────────────────────────────────────────

@app.route('/self-register/<token>', methods=['GET'])
def self_register_get(token):
    """Return student info for a valid, unused, unexpired token."""
    info = database.get_token_info(token)
    if not info:
        return jsonify({'error': 'Invalid registration link'}), 404
    if info['used_at']:
        return jsonify({'error': 'This registration link has already been used'}), 410
    from datetime import datetime as dt
    if dt.fromisoformat(info['expires_at']) < dt.now():
        return jsonify({'error': 'This registration link has expired. Please contact faculty.'}), 410
    return jsonify({
        'valid': True,
        'name':    info['name'],
        'roll_no': info['roll_no'],
        'section': info['section'],
        'course':  info['course'],
        'dept':    info['dept'],
        'room_no': info['room_no'],
        'already_registered': bool(info['has_embedding'])
    })


@app.route('/self-register/<token>', methods=['POST'])
def self_register_post(token):
    """Accept face photos from student, generate embeddings, complete registration."""
    info = database.get_token_info(token)
    if not info:
        return jsonify({'error': 'Invalid registration link'}), 404
    if info['used_at']:
        return jsonify({'error': 'This registration link has already been used'}), 410
    from datetime import datetime as dt
    if dt.fromisoformat(info['expires_at']) < dt.now():
        return jsonify({'error': 'Registration link expired'}), 410

    images = request.files.getlist('images')
    if not images or len(images) == 0:
        return jsonify({'error': 'Please provide at least one photo'}), 400

    student_id = info['student_id']
    name       = info['name']
    roll_no    = info['roll_no']

    all_embeddings = []
    failed_images  = 0

    for idx, image_file in enumerate(images[:8]):
        if not image_file or not allowed_file(image_file.filename):
            failed_images += 1
            continue
        filename = f"selfreg_{roll_no}_{idx}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.jpg"
        filepath = save_image(image_file, filename)
        try:
            faces = face_detector.detect_faces(filepath)
            if not faces:
                failed_images += 1
                logger.warning(f"[self-register] {name} image {idx}: no face detected")
                continue
            largest = max(faces, key=lambda f: (f['bbox'][2] - f['bbox'][0]) * (f['bbox'][3] - f['bbox'][1]))
            emb = face_recognizer.get_embedding(filepath, largest)
            if emb is not None:
                all_embeddings.append(emb)
            else:
                failed_images += 1
        except Exception as e:
            failed_images += 1
            logger.warning(f"[self-register] {name} image {idx} error: {e}")
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)

    if not all_embeddings:
        return jsonify({'error': 'No face could be detected in your photos. Please try with clearer, well-lit photos.'}), 400

    avg_embedding = np.mean(all_embeddings, axis=0)
    database.update_student_embedding(student_id, avg_embedding)
    for emb in all_embeddings:
        database.add_embedding(student_id, emb)
    database.mark_token_used(token)

    logger.info(f"[self-register] ✅ {name} ({roll_no}) registered — {len(all_embeddings)} embeddings")
    return jsonify({
        'success': True,
        'name': name,
        'roll_no': roll_no,
        'photos_used': len(all_embeddings),
        'photos_failed': failed_images
    })


@app.route('/students/<int:student_id>', methods=['DELETE'])
def delete_student(student_id):
    """Delete a student and their embeddings"""
    try:
        success = database.delete_student(student_id)
        if success:
            return jsonify({'success': True, 'message': 'Student deleted successfully'})
        else:
            return jsonify({'error': 'Student not found'}), 404
    except Exception as e:
        logger.error(f"Error deleting student: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/attendance/clear-today', methods=['POST'])
def clear_today_attendance():
    """Clear all attendance records for today"""
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        success = database.clear_attendance_by_date(today)
        
        if success:
            logger.info(f"Cleared all attendance records for {today}")
            return jsonify({
                'success': True, 
                'message': f'All attendance records for {today} have been cleared',
                'date': today
            })
        else:
            return jsonify({'error': 'Failed to clear attendance records'}), 500
            
    except Exception as e:
        logger.error(f"Error clearing today's attendance: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/students/clear-all', methods=['DELETE'])
def clear_all_students():
    """Delete ALL students and their embeddings - DESTRUCTIVE!"""
    try:
        success = database.clear_all_students()
        
        if success:
            logger.warning("⚠️ ALL STUDENTS AND EMBEDDINGS DELETED!")
            return jsonify({
                'success': True,
                'message': 'All students and embeddings have been deleted'
            })
        else:
            return jsonify({'error': 'Failed to clear all data'}), 500
            
    except Exception as e:
        logger.error(f"Error clearing all students: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/recognize/face', methods=['POST'])
def recognize_face():
    """
    OPTIMIZED: Recognize a single face from base64 image data - FAST VERSION
    Expected JSON: {"image": "base64_image_string"}
    Returns: {"name": "Student Name"} or {"name": null} if unknown
    """
    try:
        data = request.get_json()
        
        if not data or 'image' not in data:
            return jsonify({'error': 'No image data provided'}), 400
        
        start_time = datetime.now()
            
        # Decode base64 image - OPTIMIZED: Direct to numpy array
        base64_image = data['image']
        image_data = base64.b64decode(base64_image)
        
        # SPEED: Skip PIL, decode directly to numpy array
        nparr = np.frombuffer(image_data, np.uint8)
        image_cv = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if image_cv is None:
            logger.error("Failed to decode image")
            return jsonify({'name': None})
        
        # SPEED: Skip face detection, assume the frame already contains a cropped face
        # This is faster since Java already detected the face
        logger.info("⚡ FAST MODE: Extracting embedding directly from frame")
        
        # Create a fake face object with full image bounds
        h, w = image_cv.shape[:2]
        fake_face = {
            'bbox': [0, 0, w, h],
            'confidence': 1.0,
            'landmarks': {}
        }
        
        # Extract face embedding DIRECTLY
        embedding = face_recognizer.get_embedding(image_cv, fake_face)
        
        if embedding is None:
            logger.warning("Failed to extract embedding from face")
            return jsonify({'name': None})
        
        decode_time = (datetime.now() - start_time).total_seconds()
        logger.info(f"⏱️ Image decode + embedding: {decode_time*1000:.0f}ms")

        students = database.get_all_students()
        if not students:
            logger.info("No students registered in database")
            return jsonify({'name': None})

        match_start = datetime.now()

        # Matrix search — single BLAS multiply across all students
        result = face_recognizer.find_match(
            embedding, students, threshold=CONFIDENCE_THRESHOLD, db=database
        )

        match_time = (datetime.now() - match_start).total_seconds()
        total_time = (datetime.now() - start_time).total_seconds()

        if result is not None:
            student_id, confidence = result
            best_match = next((s for s in students if s['id'] == student_id), None)
            best_similarity = confidence
        else:
            best_match = None
            best_similarity = 0.0

        if best_match:
            logger.info(f"✅ RECOGNIZED: {best_match['name']} ({best_similarity:.4f}) in {total_time*1000:.0f}ms")
            today = datetime.now().strftime('%Y-%m-%d')
            attendance_marked = database.mark_attendance(
                best_match['id'], today, 'present', confidence=best_similarity
            )
            return jsonify({
                'name': best_match['name'],
                'student_id': best_match['id'],
                'roll_no': best_match.get('roll_no', ''),
                'similarity': best_similarity,
                'attendance_marked': attendance_marked,
                'processing_time_ms': int(total_time * 1000)
            })
        else:
            logger.info(f"❌ Not recognized (best: {best_similarity:.4f}) in {total_time*1000:.0f}ms")
            return jsonify({'name': None})
            
    except Exception as e:
        logger.error(f"Error in face recognition: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/attendance/mark', methods=['POST'])
def mark_attendance():
    """
    Manually mark a student's attendance (present or absent)
    Expected JSON: {"student_id": int, "status": "present" or "absent"}
    """
    try:
        data = request.get_json()
        
        if not data or 'student_id' not in data or 'status' not in data:
            return jsonify({'error': 'student_id and status are required'}), 400
        
        student_id = data['student_id']
        status = data['status'].lower()
        
        if status not in ['present', 'absent']:
            return jsonify({'error': 'status must be "present" or "absent"'}), 400
        
        # Get today's date
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Mark attendance in database with force_override=True for manual marking
        success = database.mark_attendance(student_id, today, status, force_override=True)
        
        if success:
            # Get student info for logging
            student = database.get_student_by_id(student_id)
            student_name = student['name'] if student else f"ID {student_id}"
            
            logger.info(f"Manually marked {student_name} as {status} for {today}")
            
            return jsonify({
                'success': True,
                'message': f'Student marked as {status}',
                'student_id': student_id,
                'status': status,
                'date': today
            })
        else:
            return jsonify({'error': 'Failed to mark attendance'}), 500
            
    except Exception as e:
        logger.error(f"Error marking attendance: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/extract_faces', methods=['POST'])
def extract_faces():
    """
    Extract all faces from a photo and return them as base64 images
    Used for adding embeddings to existing students
    """
    try:
        # Get uploaded image
        if 'image' not in request.files:
            return jsonify({'error': 'No image provided'}), 400
        
        image_file = request.files['image']
        
        if not image_file or not allowed_file(image_file.filename):
            return jsonify({'error': 'Invalid image file'}), 400
        
        logger.info("Extracting faces from uploaded image")
        
        # Save image temporarily
        filename = f"extract_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
        filepath = save_image(image_file, filename)
        
        # Detect all faces in the image
        faces = face_detector.detect_faces(filepath)
        logger.info(f"Detected {len(faces)} faces")
        
        if len(faces) == 0:
            os.remove(filepath)
            return jsonify({
                'faces': [],
                'message': 'No faces detected in the image'
            })
        
        # Read original image
        img = cv2.imread(filepath)
        
        # Extract and encode each face
        face_data_list = []
        for idx, face in enumerate(faces):
            # Extract face region
            bbox = face['bbox']
            x1, y1, x2, y2 = [int(coord) for coord in bbox]
            
            # Add margin
            margin = 30
            height, width = img.shape[:2]
            x1 = max(0, x1 - margin)
            y1 = max(0, y1 - margin)
            x2 = min(width, x2 + margin)
            y2 = min(height, y2 + margin)
            
            # Crop face
            face_img = img[y1:y2, x1:x2]
            
            # Save face temporarily
            face_filename = f"face_{idx}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
            face_path = os.path.join(UPLOAD_FOLDER, face_filename)
            cv2.imwrite(face_path, face_img)
            
            # Convert to base64
            _, buffer = cv2.imencode('.jpg', face_img)
            face_base64 = base64.b64encode(buffer).decode('utf-8')
            
            face_data_list.append({
                'index': int(idx),
                'bbox': [float(coord) for coord in bbox],  # Convert numpy types to Python float
                'confidence': float(face.get('confidence', 0.99)),  # Convert to Python float
                'image_base64': f"data:image/jpeg;base64,{face_base64}",
                'face_filename': face_filename
            })
        
        # Clean up original image
        os.remove(filepath)
        
        logger.info(f"Extracted {len(face_data_list)} faces successfully")
        
        return jsonify({
            'faces': face_data_list,
            'total_faces': len(face_data_list),
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        logger.error(f"Error extracting faces: {str(e)}", exc_info=True)
        return jsonify({'error': f'Face extraction failed: {str(e)}'}), 500

@app.route('/add_embedding/<int:student_id>', methods=['POST'])
def add_embedding_to_student(student_id):
    """
    Add new face embeddings to an existing student.
    Accepts multipart/form-data with one or more image files under the key 'images'.
    """
    try:
        student = database.get_student_by_id(student_id)
        if not student:
            return jsonify({'error': f'Student with ID {student_id} not found'}), 404

        images = request.files.getlist('images')
        if not images or len(images) == 0:
            return jsonify({'error': 'No images provided. Send image files under the key "images".'}), 400

        logger.info(f"Adding embeddings to student {student_id} ({student['name']}) — {len(images)} image(s)")

        all_embeddings = []
        failed = 0

        for idx, image_file in enumerate(images):
            if not image_file or not allowed_file(image_file.filename):
                failed += 1
                continue

            filename = f"addemb_{student_id}_{idx}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.jpg"
            filepath = save_image(image_file, filename)
            try:
                faces = face_detector.detect_faces(filepath)
                if not faces:
                    failed += 1
                    logger.warning(f"No face in image {idx} for student {student_id}")
                    continue
                largest = max(faces, key=lambda f: (f['bbox'][2] - f['bbox'][0]) * (f['bbox'][3] - f['bbox'][1]))
                embedding = face_recognizer.get_embedding(filepath, largest)
                if embedding is not None:
                    all_embeddings.append(embedding)
                else:
                    failed += 1
            except Exception as e:
                failed += 1
                logger.warning(f"Error processing image {idx} for student {student_id}: {e}")
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)

        if not all_embeddings:
            return jsonify({'error': 'No usable face found in any of the provided images'}), 400

        for emb in all_embeddings:
            database.add_embedding(student_id, emb)

        # Refresh in-memory face index
        face_recognizer.build_student_index(database.get_all_students(), database)

        logger.info(f"Added {len(all_embeddings)} embedding(s) to student {student_id} ({student['name']})")

        return jsonify({
            'success': True,
            'message': f"Added {len(all_embeddings)} embedding(s) to {student['name']}",
            'student_id': student_id,
            'student_name': student['name'],
            'embeddings_added': len(all_embeddings),
            'images_failed': failed
        })

    except Exception as e:
        logger.error(f"Error adding embedding: {str(e)}", exc_info=True)
        return jsonify({'error': f'Failed to add embedding: {str(e)}'}), 500

@app.route('/assign_face/<int:student_id>', methods=['POST'])
def assign_face_to_student(student_id):
    """
    Assign an already-extracted face crop (from /extract_faces) to a student.
    Accepts JSON: {"face_filename": "face_0_....jpg"}
    """
    try:
        student = database.get_student_by_id(student_id)
        if not student:
            return jsonify({'error': f'Student {student_id} not found'}), 404

        data = request.get_json()
        if not data or 'face_filename' not in data:
            return jsonify({'error': 'face_filename is required'}), 400

        face_filename = data['face_filename']
        face_path = os.path.join(UPLOAD_FOLDER, face_filename)

        if not os.path.exists(face_path):
            return jsonify({'error': 'Face file not found. It may have expired.'}), 404

        # Detect the face in the already-cropped image and compute its embedding
        faces = face_detector.detect_faces(face_path)
        if faces:
            largest = max(faces, key=lambda f: (f['bbox'][2] - f['bbox'][0]) * (f['bbox'][3] - f['bbox'][1]))
            embedding = face_recognizer.get_embedding(face_path, largest)
        else:
            # Crop is very tight; try with a dummy full-image bbox
            h, w = cv2.imread(face_path).shape[:2]
            dummy_face = {'bbox': [0, 0, w, h], 'confidence': 1.0}
            embedding = face_recognizer.get_embedding(face_path, dummy_face)

        if embedding is None:
            return jsonify({'error': 'Could not compute embedding for this face'}), 400

        database.add_embedding(student_id, embedding)
        face_recognizer.build_student_index(database.get_all_students(), database)

        # Clean up the temporary face crop
        try:
            os.remove(face_path)
        except Exception:
            pass

        logger.info(f"Assigned face '{face_filename}' to student {student_id} ({student['name']})")

        return jsonify({
            'success': True,
            'message': f"Face assigned to {student['name']}",
            'student_id': student_id,
            'student_name': student['name']
        })

    except Exception as e:
        logger.error(f"Error assigning face: {str(e)}", exc_info=True)
        return jsonify({'error': f'Failed to assign face: {str(e)}'}), 500


if __name__ == '__main__':
    logger.info("Starting Face Recognition Attendance System Backend...")
    logger.info(f"Confidence threshold: {CONFIDENCE_THRESHOLD}")
    app.run(host='0.0.0.0', port=5000, debug=False)  # Disable debug mode to avoid reloader issues
